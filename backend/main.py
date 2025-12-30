from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from brands.yageo.yageo_gen import generate_substitutions

app = FastAPI()

class BatchRequest(BaseModel):
    brand: str
    mpns: List[str]

# Add CORS middleware to allow frontend requests
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],  # Next.js default port
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/health")
def read_health():
    return {"status": "Healthy"}

@app.get("/api/generate")
def generate_part_substitutions(
    brand: str = Query(..., description="Brand name (e.g., yageo)"),
    mpn: str = Query(..., description="Manufacturer Part Number")
):
    """
    Generate substitutions for a given part number.
    Currently supports Yageo parts.
    """
    if brand.lower() != "yageo":
        return {
            "error": "Only Yageo brand is currently supported",
            "series": "UNKNOWN",
            "substitutions": []
        }
    
    result = generate_substitutions(mpn)
    
    return {
        "brand": brand,
        "mpn": mpn,
        "series": result["series"],
        "substitutions": result["substitutions"]
    }

@app.post("/api/generate/batch")
def generate_batch_substitutions(request: BatchRequest):
    """
    Generate substitutions for multiple part numbers in parallel.
    Currently supports Yageo parts.
    """
    if request.brand.lower() != "yageo":
        return {
            "error": "Only Yageo brand is currently supported",
            "brand": request.brand,
            "total": 0,
            "results": []
        }
    
    def process_single_mpn(mpn: str):
        result = generate_substitutions(mpn)
        return {
            "mpn": mpn,
            "series": result["series"],
            "substitutions": result["substitutions"]
        }
    
    # Process MPNs in parallel using ThreadPoolExecutor
    with ThreadPoolExecutor(max_workers=20) as executor:
        results = list(executor.map(process_single_mpn, request.mpns))
    
    return {
        "brand": request.brand,
        "total": len(results),
        "results": results
    }

@app.post("/api/generate/batch/export")
def export_batch_to_excel(request: BatchRequest):
    """
    Generate substitutions for multiple part numbers and export to Excel.
    Currently supports Yageo parts.
    """
    if request.brand.lower() != "yageo":
        return {
            "error": "Only Yageo brand is currently supported"
        }
    
    def process_single_mpn(mpn: str):
        result = generate_substitutions(mpn)
        return {
            "mpn": mpn,
            "series": result["series"],
            "substitutions": result["substitutions"]
        }
    
    # Process MPNs in parallel
    with ThreadPoolExecutor(max_workers=20) as executor:
        results = list(executor.map(process_single_mpn, request.mpns))
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Substitutions"
    
    # Define header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="left", vertical="center")
    
    # Set headers
    headers = ["MPN", "Substitution", "Type", "Details"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Set column widths
    ws.column_dimensions['A'].width = 25  # MPN
    ws.column_dimensions['B'].width = 25  # Substitution
    ws.column_dimensions['C'].width = 25  # Type
    ws.column_dimensions['D'].width = 50  # Details
    
    # Add data - one row per substitution
    row_num = 2
    for result in results:
        mpn = result["mpn"]
        for sub in result["substitutions"]:
            ws.cell(row=row_num, column=1, value=mpn)
            ws.cell(row=row_num, column=2, value=sub["part_number"])
            ws.cell(row=row_num, column=3, value=sub["type"])
            ws.cell(row=row_num, column=4, value=sub["details"])
            row_num += 1
    
    # Save to bytes
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename with timestamp
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"yageo_substitutions_{timestamp}.xlsx"
    
    # Return as downloadable file
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
    )